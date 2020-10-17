from PyQt5 import QtCore, QtGui, QtWidgets
import logo_rc
from trial2 import  Ui_SuccessDialog
import xlrd
import openpyxl #to load and open excel sheetand workbook
import mysql.connector
from mysql.connector import Error
from PyQt5.QtWidgets import QMessageBox
from datetime import date
import os
import time

file_name = ""

dict = {'bd':127,'nb':127,'bh':5,'ew':250} #dictionary to store type and cost of weight


class Ui_MainWindow(object):

    def graph_fetch_handler(self): 
            print("plot graph")

    def cost_fetch_handler(self):
            global file_name
            msg = QMessageBox()
            msg.setText("Please select a file first")
            connection = mysql.connector.connect(host='localhost',database='waste_segregator',user='root',password='',buffered=True)
            cur5 = connection.cursor()
            select_cost_qry =  "SELECT waste_cost FROM waste inner join files on files.file_id = waste.file_id_fk WHERE files.file_link=\"{x}\"".format(x=file_name)
            cur5.execute(select_cost_qry)
            if cur5.rowcount == 0:
                    msg.exec_()
                    connection.close()
            else:

                results = cur5.fetchall()
                print(results)
                self.label_22.setText(str(results[0][0]))
                self.label_34.setText(str(results[1][0]))
                self.label_30.setText(str(results[2][0]))
                self.label_32.setText(str(results[3][0]))
                connection.close()

    def weight_fetch_handler(self):
            global file_name
            msg = QMessageBox()
            msg.setText("Please select a file first")
            connection = mysql.connector.connect(host='localhost',
                                                database='waste_segregator',
                                                user='root',
                                                password='',buffered=True)
            cur4 = connection.cursor()
            select_weight_qry =  "SELECT waste_weight FROM waste inner join files on files.file_id = waste.file_id_fk WHERE files.file_link=\"{x}\"".format(x=file_name)
            cur4.execute(select_weight_qry)
            if cur4.rowcount == 0:
                    msg.exec_()
                    connection.close()
            else:

                results = cur4.fetchall()
                print(results)
                self.bd_qty_label.setText(str(results[0][0]))
                self.nb_qty_label.setText(str(results[1][0]))
                self.bh_qty_label.setText(str(results[2][0]))
                self.ew_qty_label.setText(str(results[3][0]))
                

                connection.close()
  
    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(964, 780)
        icon = QtGui.QIcon()
        icon.addPixmap(QtGui.QPixmap(":/images/logo-check-1.JPG"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        MainWindow.setWindowIcon(icon)
        MainWindow.setStyleSheet("")
        MainWindow.setIconSize(QtCore.QSize(55, 50))
        MainWindow.setTabShape(QtWidgets.QTabWidget.Rounded)
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.horizontalLayout = QtWidgets.QHBoxLayout(self.centralwidget)
        self.horizontalLayout.setObjectName("horizontalLayout")
        self.tabWidget = QtWidgets.QTabWidget(self.centralwidget)
        self.tabWidget.setStyleSheet("font: 18pt \"Calibri\";\n"
"background-color: rgb(179, 224, 242);\n"
"")
        self.tabWidget.setTabShape(QtWidgets.QTabWidget.Rounded)
        self.tabWidget.setIconSize(QtCore.QSize(55, 30))
        self.tabWidget.setObjectName("tabWidget")
        self.uploadtab = QtWidgets.QWidget()
        self.uploadtab.setObjectName("uploadtab")
        self.calculateButton = QtWidgets.QPushButton(self.uploadtab)
        self.calculateButton.setGeometry(QtCore.QRect(330, 360, 221, 51))
        self.calculateButton.setStyleSheet("background-color: rgb(113, 177, 217);\n"
"color: rgb(255, 255, 255);\n"
"font: 18pt \"Calibri\";\n"
"border-radius: 15px;\n"
"\n"
"")
        self.calculateButton.setObjectName("calculateButton")
        self.calculateButton.clicked.connect(submit_button_handler)
        self.browseButton = QtWidgets.QPushButton(self.uploadtab)
        self.browseButton.setGeometry(QtCore.QRect(520, 180, 161, 31))
        '''self.browseButton.setStyleSheet("font: 14pt \"Calibri\";\n"
"color: rgb(255, 255, 255);\n"
"border-radius:15px;\n"
"background-color: #8c8880;\n"
"")'''
        self.browseButton.setStyleSheet("QPushButton::hover"
                             "{"
                             "background-color : lightgreen;"
                             "}\n"
                             "font: 14pt \"Calibri\";\n"
"color: rgb(255, 255, 255);\n"
"border-radius:15px;\n"
"background-color: #8c8880;\n"
"")
        self.browseButton.setObjectName("browseButton")
        self.browseButton.clicked.connect(browse_button_handler)
        self.upload_image_label = QtWidgets.QLabel(self.uploadtab)
        self.upload_image_label.setGeometry(QtCore.QRect(260, 130, 111, 131))
        self.upload_image_label.setFrameShape(QtWidgets.QFrame.NoFrame)
        self.upload_image_label.setText("")
        self.upload_image_label.setPixmap(QtGui.QPixmap(":/images/files.png"))
        self.upload_image_label.setScaledContents(True)
        self.upload_image_label.setAlignment(QtCore.Qt.AlignCenter)
        self.upload_image_label.setObjectName("upload_image_label")
        self.instruction_label = QtWidgets.QLabel(self.uploadtab)
        self.instruction_label.setGeometry(QtCore.QRect(20, 460, 381, 81))
        self.instruction_label.setStyleSheet("font: 25 10pt \"Calibri Light\";")
        self.instruction_label.setObjectName("instruction_label")
        icon1 = QtGui.QIcon()
        icon1.addPixmap(QtGui.QPixmap(":/images/upload-icon.png"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.tabWidget.addTab(self.uploadtab, icon1, "")
        self.resulttab = QtWidgets.QWidget()
        self.resulttab.setObjectName("resulttab")
        

        #adding the fetch-data button here
        self.weightFetchButton = QtWidgets.QPushButton(self.resulttab)
        self.weightFetchButton.setGeometry(QtCore.QRect(750, 50, 161, 31))
        self.weightFetchButton.setStyleSheet("QPushButton::hover"
                             "{"
                             "background-color : lightgreen;"
                             "}\n"
                             "font: 12pt \"Calibri\";\n"
"color: rgb(255, 255, 255);\n"
"border-radius:15px;\n"
"background-color: #ffffff;\n"
"")
        self.weightFetchButton.setObjectName("weightFetchButton")
        self.weightFetchButton.setText("Fetch Data")
        self.weightFetchButton.clicked.connect(self.weight_fetch_handler)


        self.nb_waste_image_label = QtWidgets.QLabel(self.resulttab)
        self.nb_waste_image_label.setGeometry(QtCore.QRect(510, 30, 201, 141))
        self.nb_waste_image_label.setAutoFillBackground(False)
        self.nb_waste_image_label.setText("")
        self.nb_waste_image_label.setPixmap(QtGui.QPixmap(":/images/non-bio.jpeg"))
        self.nb_waste_image_label.setScaledContents(True)
        self.nb_waste_image_label.setObjectName("nb_waste_image_label")
        self.verticalLayoutWidget_2 = QtWidgets.QWidget(self.resulttab)
        self.verticalLayoutWidget_2.setGeometry(QtCore.QRect(510, 190, 211, 154))
        self.verticalLayoutWidget_2.setObjectName("verticalLayoutWidget_2")
        self.verticalLayout_2 = QtWidgets.QVBoxLayout(self.verticalLayoutWidget_2)
        self.verticalLayout_2.setContentsMargins(0, 0, 0, 0)
        self.verticalLayout_2.setObjectName("verticalLayout_2")
        self.nb_waste_label = QtWidgets.QLabel(self.verticalLayoutWidget_2)
        self.nb_waste_label.setStyleSheet("font: 14pt \"Calibri\";")
        self.nb_waste_label.setAlignment(QtCore.Qt.AlignCenter)
        self.nb_waste_label.setObjectName("nb_waste_label")
        self.verticalLayout_2.addWidget(self.nb_waste_label)
        self.horizontalLayout_2 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_2.setObjectName("horizontalLayout_2")
        self.nb_qty_label = QtWidgets.QLabel(self.verticalLayoutWidget_2)
        self.nb_qty_label.setStyleSheet("font: 75 10pt \"Calibri\";")
        self.nb_qty_label.setAlignment(QtCore.Qt.AlignCenter)
        self.nb_qty_label.setObjectName("nb_qty_label")
        self.horizontalLayout_2.addWidget(self.nb_qty_label)
        
        
        self.verticalLayout_2.addLayout(self.horizontalLayout_2)
        self.verticalLayoutWidget_3 = QtWidgets.QWidget(self.resulttab)
        self.verticalLayoutWidget_3.setGeometry(QtCore.QRect(160, 490, 211, 128))
        self.verticalLayoutWidget_3.setObjectName("verticalLayoutWidget_3")
        self.verticalLayout_3 = QtWidgets.QVBoxLayout(self.verticalLayoutWidget_3)
        self.verticalLayout_3.setContentsMargins(0, 0, 0, 0)
        self.verticalLayout_3.setObjectName("verticalLayout_3")
        self.label_12 = QtWidgets.QLabel(self.verticalLayoutWidget_3)
        self.label_12.setStyleSheet("font: 14pt \"Calibri\";")
        self.label_12.setAlignment(QtCore.Qt.AlignCenter)
        self.label_12.setObjectName("label_12")
        self.verticalLayout_3.addWidget(self.label_12)
        self.horizontalLayout_3 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_3.setObjectName("horizontalLayout_3")
        self.bh_qty_label = QtWidgets.QLabel(self.verticalLayoutWidget_3)
        self.bh_qty_label.setStyleSheet("font: 75 10pt \"Calibri\";")
        self.bh_qty_label.setAlignment(QtCore.Qt.AlignCenter)
        self.bh_qty_label.setObjectName("bh_qty_label")
        self.horizontalLayout_3.addWidget(self.bh_qty_label)

        self.verticalLayout_3.addLayout(self.horizontalLayout_3)
        self.ew_image_label = QtWidgets.QLabel(self.resulttab)
        self.ew_image_label.setGeometry(QtCore.QRect(520, 360, 201, 111))
        self.ew_image_label.setAutoFillBackground(False)
        self.ew_image_label.setText("")
        self.ew_image_label.setPixmap(QtGui.QPixmap(":/images/e-waste.jpeg"))
        self.ew_image_label.setScaledContents(True)
        self.ew_image_label.setObjectName("ew_image_label")
        self.bd_waste_image_label = QtWidgets.QLabel(self.resulttab)
        self.bd_waste_image_label.setGeometry(QtCore.QRect(140, 40, 201, 141))
        self.bd_waste_image_label.setAutoFillBackground(False)
        self.bd_waste_image_label.setStyleSheet("image: url(:/images/bio-waste.jpg);")
        self.bd_waste_image_label.setText("")
        self.bd_waste_image_label.setPixmap(QtGui.QPixmap(":/images/bio-waste.jpg"))
        self.bd_waste_image_label.setScaledContents(True)
        self.bd_waste_image_label.setObjectName("bd_waste_image_label")
        self.bh_waste_image_label = QtWidgets.QLabel(self.resulttab)
        self.bh_waste_image_label.setGeometry(QtCore.QRect(160, 360, 201, 111))
        self.bh_waste_image_label.setAutoFillBackground(False)
        self.bh_waste_image_label.setText("")
        self.bh_waste_image_label.setPixmap(QtGui.QPixmap(":/images/Biohazard_symbol_(black_and_yellow).png"))####issue with image
        self.bh_waste_image_label.setScaledContents(True)
        self.bh_waste_image_label.setObjectName("bh_waste_image_label")
        self.verticalLayoutWidget = QtWidgets.QWidget(self.resulttab)
        self.verticalLayoutWidget.setGeometry(QtCore.QRect(520, 490, 201, 128))
        self.verticalLayoutWidget.setObjectName("verticalLayoutWidget")
        self.verticalLayout = QtWidgets.QVBoxLayout(self.verticalLayoutWidget)
        self.verticalLayout.setContentsMargins(0, 0, 0, 0)
        self.verticalLayout.setObjectName("verticalLayout")
        self.label_13 = QtWidgets.QLabel(self.verticalLayoutWidget)
        self.label_13.setStyleSheet("font: 14pt \"Calibri\";")
        self.label_13.setAlignment(QtCore.Qt.AlignCenter)
        self.label_13.setObjectName("label_13")
        self.verticalLayout.addWidget(self.label_13)
        self.horizontalLayout_4 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_4.setObjectName("horizontalLayout_4")
        self.ew_qty_label = QtWidgets.QLabel(self.verticalLayoutWidget)
        self.ew_qty_label.setStyleSheet("font: 75 10pt \"Calibri\";")
        self.ew_qty_label.setAlignment(QtCore.Qt.AlignCenter)
        self.ew_qty_label.setObjectName("ew_qty_label")
        self.horizontalLayout_4.addWidget(self.ew_qty_label)

        self.verticalLayout.addLayout(self.horizontalLayout_4)
        self.verticalLayoutWidget_4 = QtWidgets.QWidget(self.resulttab)
        self.verticalLayoutWidget_4.setGeometry(QtCore.QRect(140, 200, 201, 128))
        self.verticalLayoutWidget_4.setObjectName("verticalLayoutWidget_4")
        self.verticalLayout_4 = QtWidgets.QVBoxLayout(self.verticalLayoutWidget_4)
        self.verticalLayout_4.setContentsMargins(0, 0, 0, 0)
        self.verticalLayout_4.setObjectName("verticalLayout_4")
        self.bd_waste_label = QtWidgets.QLabel(self.verticalLayoutWidget_4)
        self.bd_waste_label.setStyleSheet("font: 14pt \"Calibri\";")
        self.bd_waste_label.setAlignment(QtCore.Qt.AlignCenter)
        self.bd_waste_label.setObjectName("bd_waste_label")
        self.verticalLayout_4.addWidget(self.bd_waste_label)
        self.horizontalLayout_5 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_5.setObjectName("horizontalLayout_5")
        self.bd_qty_label = QtWidgets.QLabel(self.verticalLayoutWidget_4)
        self.bd_qty_label.setStyleSheet("font: 75 10pt \"Calibri\";")
        self.bd_qty_label.setAlignment(QtCore.Qt.AlignCenter)
        self.bd_qty_label.setObjectName("bd_qty_label")
        self.horizontalLayout_5.addWidget(self.bd_qty_label)

        self.verticalLayout_4.addLayout(self.horizontalLayout_5)
        icon2 = QtGui.QIcon()
        icon2.addPixmap(QtGui.QPixmap(":/images/baseline_build_black_18dp.png"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.tabWidget.addTab(self.resulttab, icon2, "")
        self.statstab = QtWidgets.QWidget()
        self.statstab.setObjectName("statstab")

        self.graphFetchButton = QtWidgets.QPushButton(self.statstab)
        self.graphFetchButton.setGeometry(QtCore.QRect(750, 50, 161, 31))
        self.graphFetchButton.setStyleSheet("QPushButton::hover"
                             "{"
                             "background-color : lightgreen;"
                             "}\n"
                             "font: 12pt \"Calibri\";\n"
"color: rgb(255, 255, 255);\n"
"border-radius:15px;\n"
"background-color: #ffffff;\n"
"")
        self.graphFetchButton.setObjectName("graph")
        self.graphFetchButton.setText("Plot Data")
        self.graphFetchButton.clicked.connect(self.graph_fetch_handler)

        icon3 = QtGui.QIcon()
        icon3.addPixmap(QtGui.QPixmap(":/images/chart.png"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.tabWidget.addTab(self.statstab, icon3, "")
        self.costtab = QtWidgets.QWidget()
        self.costtab.setObjectName("costtab")
        #adding the fetch-data-weight button here
        self.costFetchButton = QtWidgets.QPushButton(self.costtab)
        self.costFetchButton.setGeometry(QtCore.QRect(750, 50, 161, 31))
        self.costFetchButton.setStyleSheet("QPushButton::hover"
                             "{"
                             "background-color : lightgreen;"
                             "}\n"
                             "font: 12pt \"Calibri\";\n"
"color: rgb(255, 255, 255);\n"
"border-radius:15px;\n"
"background-color: #ffffff;\n"
"")
        self.costFetchButton.setObjectName("costFetchButton")
        self.costFetchButton.setText("Fetch Cost")
        self.costFetchButton.clicked.connect(self.cost_fetch_handler)


        self.stats_heading_label = QtWidgets.QLabel(self.costtab)
        self.stats_heading_label.setGeometry(QtCore.QRect(310, 20, 191, 51))
        self.stats_heading_label.setStyleSheet("font: 22pt \"Calibri\";")
        self.stats_heading_label.setObjectName("stats_heading_label")
        self.bd_image_label = QtWidgets.QLabel(self.costtab)
        self.bd_image_label.setGeometry(QtCore.QRect(120, 70, 201, 141))
        self.bd_image_label.setText("")
        self.bd_image_label.setPixmap(QtGui.QPixmap(":/images/bio-waste.jpg"))
        self.bd_image_label.setScaledContents(True)
        self.bd_image_label.setObjectName("bd_image_label")
        self.nb_img_label = QtWidgets.QLabel(self.costtab)
        self.nb_img_label.setGeometry(QtCore.QRect(540, 60, 201, 141))
        self.nb_img_label.setText("")
        self.nb_img_label.setPixmap(QtGui.QPixmap(":/images/non-bio.jpeg"))
        self.nb_img_label.setScaledContents(True)
        self.nb_img_label.setObjectName("nb_img_label")
        self.bh_img_label = QtWidgets.QLabel(self.costtab)
        self.bh_img_label.setGeometry(QtCore.QRect(120, 380, 201, 141))
        self.bh_img_label.setText("")
        self.bh_img_label.setPixmap(QtGui.QPixmap(":/images/Biohazard_symbol_(black_and_yellow).png"))
        self.bh_img_label.setScaledContents(True)
        self.bh_img_label.setObjectName("bh_img_label")
        self.ew_img_label = QtWidgets.QLabel(self.costtab)
        self.ew_img_label.setGeometry(QtCore.QRect(540, 380, 201, 141))
        self.ew_img_label.setText("")
        self.ew_img_label.setPixmap(QtGui.QPixmap(":/images/e-waste.jpeg"))
        self.ew_img_label.setScaledContents(True)
        self.ew_img_label.setObjectName("ew_img_label")
        self.verticalLayoutWidget_5 = QtWidgets.QWidget(self.costtab)
        self.verticalLayoutWidget_5.setGeometry(QtCore.QRect(100, 230, 231, 155))
        self.verticalLayoutWidget_5.setObjectName("verticalLayoutWidget_5")
        self.verticalLayout_5 = QtWidgets.QVBoxLayout(self.verticalLayoutWidget_5)
        self.verticalLayout_5.setContentsMargins(10, 0, 0, 10)
        self.verticalLayout_5.setObjectName("verticalLayout_5")
        self.bd_title_label = QtWidgets.QLabel(self.verticalLayoutWidget_5)
        self.bd_title_label.setStyleSheet("font: 14pt \"Calibri\";")
        self.bd_title_label.setAlignment(QtCore.Qt.AlignCenter)
        self.bd_title_label.setObjectName("bd_title_label")
        self.verticalLayout_5.addWidget(self.bd_title_label)
        self.horizontalLayout_6 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_6.setContentsMargins(10, 10, 10, 10)
        self.horizontalLayout_6.setObjectName("horizontalLayout_6")
        self.label_22 = QtWidgets.QLabel(self.verticalLayoutWidget_5)
        self.label_22.setStyleSheet("font: 12pt \"Calibri\";")
        self.label_22.setAlignment(QtCore.Qt.AlignCenter)
        self.label_22.setObjectName("label_22")
        self.horizontalLayout_6.addWidget(self.label_22)

        self.verticalLayout_5.addLayout(self.horizontalLayout_6)
        self.verticalLayoutWidget_9 = QtWidgets.QWidget(self.costtab)
        self.verticalLayoutWidget_9.setGeometry(QtCore.QRect(100, 544, 241, 155))
        self.verticalLayoutWidget_9.setObjectName("verticalLayoutWidget_9")
        self.verticalLayout_9 = QtWidgets.QVBoxLayout(self.verticalLayoutWidget_9)
        self.verticalLayout_9.setContentsMargins(10, 0, 0, 10)
        self.verticalLayout_9.setObjectName("verticalLayout_9")
        self.bh_title_label = QtWidgets.QLabel(self.verticalLayoutWidget_9)
        self.bh_title_label.setStyleSheet("font: 14pt \"Calibri\";")
        self.bh_title_label.setAlignment(QtCore.Qt.AlignCenter)
        self.bh_title_label.setObjectName("bh_title_label")
        self.verticalLayout_9.addWidget(self.bh_title_label)
        self.horizontalLayout_7 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_7.setContentsMargins(10, 10, 10, 10)
        self.horizontalLayout_7.setObjectName("horizontalLayout_7")
        self.label_30 = QtWidgets.QLabel(self.verticalLayoutWidget_9)
        self.label_30.setStyleSheet("font: 12pt \"Calibri\";")
        self.label_30.setAlignment(QtCore.Qt.AlignCenter)
        self.label_30.setObjectName("label_30")
        self.horizontalLayout_7.addWidget(self.label_30)

        self.verticalLayout_9.addLayout(self.horizontalLayout_7)
        self.verticalLayoutWidget_10 = QtWidgets.QWidget(self.costtab)
        self.verticalLayoutWidget_10.setGeometry(QtCore.QRect(520, 544, 241, 155))
        self.verticalLayoutWidget_10.setObjectName("verticalLayoutWidget_10")
        self.verticalLayout_10 = QtWidgets.QVBoxLayout(self.verticalLayoutWidget_10)
        self.verticalLayout_10.setContentsMargins(10, 0, 0, 10)
        self.verticalLayout_10.setObjectName("verticalLayout_10")
        self.ew_title_label = QtWidgets.QLabel(self.verticalLayoutWidget_10)
        self.ew_title_label.setStyleSheet("font: 14pt \"Calibri\";")
        self.ew_title_label.setAlignment(QtCore.Qt.AlignCenter)
        self.ew_title_label.setObjectName("ew_title_label")
        self.verticalLayout_10.addWidget(self.ew_title_label)
        self.horizontalLayout_8 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_8.setContentsMargins(10, 10, 10, 10)
        self.horizontalLayout_8.setObjectName("horizontalLayout_8")
        self.label_32 = QtWidgets.QLabel(self.verticalLayoutWidget_10)
        self.label_32.setStyleSheet("font: 12pt \"Calibri\";")
        self.label_32.setAlignment(QtCore.Qt.AlignCenter)
        self.label_32.setObjectName("label_32")
        self.horizontalLayout_8.addWidget(self.label_32)

        self.verticalLayout_10.addLayout(self.horizontalLayout_8)
        self.verticalLayoutWidget_11 = QtWidgets.QWidget(self.costtab)
        self.verticalLayoutWidget_11.setGeometry(QtCore.QRect(520, 220, 241, 155))
        self.verticalLayoutWidget_11.setObjectName("verticalLayoutWidget_11")
        self.verticalLayout_11 = QtWidgets.QVBoxLayout(self.verticalLayoutWidget_11)
        self.verticalLayout_11.setContentsMargins(10, 0, 0, 10)
        self.verticalLayout_11.setObjectName("verticalLayout_11")
        self.nb_title_label = QtWidgets.QLabel(self.verticalLayoutWidget_11)
        self.nb_title_label.setStyleSheet("font: 14pt \"Calibri\";")
        self.nb_title_label.setAlignment(QtCore.Qt.AlignCenter)
        self.nb_title_label.setObjectName("nb_title_label")
        self.verticalLayout_11.addWidget(self.nb_title_label)
        self.horizontalLayout_9 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_9.setContentsMargins(10, 10, 10, 10)
        self.horizontalLayout_9.setObjectName("horizontalLayout_9")
        self.label_34 = QtWidgets.QLabel(self.verticalLayoutWidget_11)
        self.label_34.setStyleSheet("font: 12pt \"Calibri\";")
        self.label_34.setAlignment(QtCore.Qt.AlignCenter)
        self.label_34.setObjectName("label_34")
        self.horizontalLayout_9.addWidget(self.label_34)
        self.verticalLayout_11.addLayout(self.horizontalLayout_9)
        icon4 = QtGui.QIcon()
        icon4.addPixmap(QtGui.QPixmap(":/images/baseline_monetization_on_black_18dp.png"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.tabWidget.addTab(self.costtab, icon4, "")
        self.horizontalLayout.addWidget(self.tabWidget)
        MainWindow.setCentralWidget(self.centralwidget)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)
        self.appGuide = QtWidgets.QAction(MainWindow)
        icon5 = QtGui.QIcon()
        icon5.addPixmap(QtGui.QPixmap(":/images/help-icon.png"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.appGuide.setIcon(icon5)
        self.appGuide.setIconVisibleInMenu(True)
        self.appGuide.setObjectName("appGuide")
        self.appTools = QtWidgets.QAction(MainWindow)
        icon6 = QtGui.QIcon()
        icon6.addPixmap(QtGui.QPixmap(":/images/tools-icon.png"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.appTools.setIcon(icon6)
        self.appTools.setObjectName("appTools")
        self.appExit = QtWidgets.QAction(MainWindow)
        self.appExit.setIconVisibleInMenu(False)
        self.appExit.setShortcutVisibleInContextMenu(True)
        self.appExit.setObjectName("appExit")

        self.retranslateUi(MainWindow)
        self.tabWidget.setCurrentIndex(0)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)
            
                   
    def retranslateUi(self, MainWindow):
        global weights
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "Wate Segregator"))
        self.calculateButton.setText(_translate("MainWindow", "SUBMIT"))
        self.browseButton.setText(_translate("MainWindow", "browse"))
        self.instruction_label.setText(_translate("MainWindow", "Upload your data file here\n"
"Instructions-File should have the following -\n"
"1. Waste weight, 2. Waste Type, 3. Waste Name 4. Date"))
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.uploadtab), _translate("MainWindow", "Upload file"))
        self.nb_waste_label.setText(_translate("MainWindow", "Non - Biodegradable\n (Quantity - kg)"))
        self.nb_qty_label.setText(_translate("MainWindow", "0"))
        self.nb_qty_label.setText(_translate("MainWindow","0"))
        self.label_12.setText(_translate("MainWindow", "Bio - hazardous\n (Quantity - kg)"))
        self.bh_qty_label.setText(_translate("MainWindow", "0"))
        self.label_13.setText(_translate("MainWindow", "E - waste\n (Quantity - kg)"))
        self.ew_qty_label.setText(_translate("MainWindow", "0 "))
        self.bd_waste_label.setText(_translate("MainWindow", "Bio-Degradable\n (Quantity - kg)"))
        self.bd_qty_label.setText(_translate("MainWindow", "0"))
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.resulttab), _translate("MainWindow", "Results"))
       # self.stats_label.setText(_translate("MainWindow", "Graphical analysis"))
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.statstab), _translate("MainWindow", "Statistics"))
        self.stats_heading_label.setText(_translate("MainWindow", "Cost analysis"))
        self.bd_title_label.setText(_translate("MainWindow", "Bio-degradable\n (Cost - Rs)"))
        self.label_22.setText(_translate("MainWindow", "0"))
        self.bh_title_label.setText(_translate("MainWindow", "Bio-hazardous\n (Cost - Rs)"))
        self.label_30.setText(_translate("MainWindow", "0"))
        self.ew_title_label.setText(_translate("MainWindow", "E-waste\n (Cost - Rs)"))
        self.label_32.setText(_translate("MainWindow", "0"))
        self.nb_title_label.setText(_translate("MainWindow", "Non - Biodegradable\n (Cost - Rs)"))
        self.label_34.setText(_translate("MainWindow", "0"))
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.costtab), _translate("MainWindow", "Cost"))
        
        self.appGuide.setText(_translate("MainWindow", "Guide"))
        self.appTools.setText(_translate("MainWindow", "Tools"))
        self.appExit.setText(_translate("MainWindow", "Exit"))
        self.appExit.setShortcut(_translate("MainWindow", "Ctrl+X"))
     

    

def browse_button_handler(self): #function for browse button handling
        
        filename = QtWidgets.QFileDialog.getOpenFileName(None,('Choose file'),"/home",("Xcel sheet(*.xlsx)")) #filename result is a tuple
        path = filename[0]
        wb = openpyxl.load_workbook(path) #loading the workbook
        sh = wb.active #loading the active sheet
        print(sh)
        global file_name
        global weights

        file_name = str(os.path.basename(path)) 
        file_names = [file_name]
        print(file_names)
        try:
                connection = mysql.connector.connect(host='localhost',
                                                database='waste_segregator',
                                                user='root',
                                                password='')

                if connection.is_connected():
                        db_Info = connection.get_server_info()
                        print("Connected to MySQL Server version ", db_Info)
                        cursor = connection.cursor()
                        #check first if the record exixts, if yes, file wont get update/inserted again

                        file_id_check_qry = "select file_link from files where file_link =\"{x}\"".format(x=file_name)
                        cursor.execute(file_id_check_qry)
                        if cursor.rowcount == 0: #file does not exist in DB add it first 
                                file_insert_qry = "insert into files (file_link) values (%s)"
                                cursor.execute(file_insert_qry,file_names)                       
                                connection.commit()
                                print("file name",file_name,"inserted!!")
                        else:
                                print("This file already exits in the DB")


        except Error as e:
                print("Error while connecting to MySQL", e)
        finally:
                if (connection.is_connected()):
                        cursor.close()
                        connection.close()
                        print("MySQL connection is closed")
        
        waste_data = [] #will contain all data of the waste
        for r in sh.iter_rows(min_row = sh.min_row + 1, max_row = sh.max_row, min_col = 1, max_col =4): 
                row_data = [] #will contain row-wise data of the sheet
                for c in r:
                        row_data.append(c.value)
                        
                waste_data.append(row_data)
        bd_cost = 0
        nb_cost = 0
        bh_cost = 0
        ew_cost = 0
        bd_weight=0
        nb_weight=0
        bh_weight=0
        ew_weight=0
        for i in waste_data:
                if(i[1] == "bd"):
                        bd_cost = bd_cost + i[0]*dict["bd"]
                        bd_weight = bd_weight + i[0]
                if(i[1] == "nb"):
                        nb_cost = nb_cost + i[0]*dict["nb"]
                        nb_weight = nb_weight + i[0]
                if(i[1] == "bh"):
                        bh_cost = bh_cost + i[0]*dict["bh"]
                        bh_weight = bh_weight + i[0]
                    
                if(i[1] == "ew"):
                        ew_cost = ew_cost + i[0]*dict["ew"]
                        ew_weight = ew_weight + i[0]
        current_date = date.today()       
        bd_entry = [bd_cost,bd_weight,"bd",current_date]
        nb_entry = [nb_cost,nb_weight,"nb",current_date]
        bh_entry = [bh_cost,bh_weight,"bh",current_date]
        ew_entry = [ew_cost,ew_weight,"ew",current_date]
        total_data = [bd_entry,nb_entry,bh_entry,ew_entry]
        print(total_data)

        try:
                connection = mysql.connector.connect(host='localhost',
                                                database='waste_segregator',
                                                user='root',
                                                password='',buffered=True)
                file_name1 = "hello"

                if connection.is_connected():
                        db_Info = connection.get_server_info()
                        print("Connected to MySQL Server version ", db_Info)
                        cursor = connection.cursor()
                        cursor1 = connection.cursor()
                        cursor3 = connection.cursor()
                        sql_select_qry = "SELECT file_id FROM files WHERE file_link=\"{x}\"".format(x=file_name)
                        sql_select_qry1 = "SELECT file_id_fk FROM waste inner join files on files.file_id = waste.file_id_fk WHERE files.file_link=\"{x}\"".format(x=file_name)
                        qry =  "SELECT waste_weight FROM waste inner join files on files.file_id = waste.file_id_fk WHERE files.file_link=\"{x}\"".format(x=file_name)

                        cursor.execute(sql_select_qry)
                        cursor1.execute(sql_select_qry1)
                        cursor3.execute(qry)
                        weights = cursor3.fetchall()
                        print(weights)
                        print("File ID now selected")
                        records = cursor.fetchall() #fetches the file_id from the tabble
                        print(total_data)
                        if cursor1.rowcount == 0:
                                sql_insert = "insert into waste (waste_cost, waste_weight, waste_type, waste_date, file_id_fk) values (%s,%s,%s,%s,{x})".format(x=records[0][0])
                                cursor.executemany(sql_insert,total_data)                       
                                connection.commit()
                        else:
                                cursor2 = connection.cursor()
                                boombaam_qry = "select waste_id from waste where file_id_fk = {x}".format(x=records[0][0]) #selecting the IDs
                                y = cursor2.execute(boombaam_qry)
                                yz = cursor2.fetchall()
                                print(yz[0])
                                print(y)
                                
                                print("File ID already in that table hence we udate and not insert")
                                for i in range(len(total_data)):
                                        print(total_data[i][3])
                                        sql_update_qry = "UPDATE waste SET waste_cost = {a}, waste_weight = {b}, waste_type = \"{c}\", waste_date = \"{d}\" where waste_id = {x}".format(x=yz[i][0],a=total_data[i][0],b=total_data[i][1],c=str(total_data[i][2]),d=total_data[i][3])
                                        #print(total_data[i])
                                        cursor.execute(sql_update_qry) 
                                        connection.commit()
                                print("Update query chalgae")
                        print(records)

        except Error as e:
                print("Error while connecting to MySQL", e)
        finally:
                if (connection.is_connected()):
                        cursor.close()
                        connection.close()
                        print("MySQL connection is closed")

def submit_button_handler(self): #function for submit button hadling #have to make a connection with database and check if rows actually got inserted or not

    SuccessDialog = QtWidgets.QDialog()
    print("hello")
    ui = Ui_SuccessDialog()
    ui.setupUi(SuccessDialog)
    SuccessDialog.show()
    SuccessDialog.exec_() #execute the dialog
#from pyqtgraph import PlotWidget


''' Start dialog class and functions'''
class Ui_StartDialog(object):
    def setupUi(self, StartDialog):
        StartDialog.setObjectName("StartDialog")
        StartDialog.resize(850, 600)
        StartDialog.setStyleSheet("background-color: rgb(179, 224, 242);")
        self.gridLayout = QtWidgets.QGridLayout(StartDialog)
        self.gridLayout.setObjectName("gridLayout")
        self.verticalLayout_2 = QtWidgets.QVBoxLayout()
        self.verticalLayout_2.setObjectName("verticalLayout_2")
        self.imagelabel = QtWidgets.QLabel(StartDialog)
        self.imagelabel.setText("")
        self.imagelabel.setPixmap(QtGui.QPixmap(":/images/logo-check-1.JPG"))
        self.imagelabel.setScaledContents(True)
        self.imagelabel.setObjectName("imagelabel")
        self.verticalLayout_2.addWidget(self.imagelabel)
        self.horizontalLayout = QtWidgets.QHBoxLayout()
        self.horizontalLayout.setContentsMargins(20, 20, 5, 5)
        self.horizontalLayout.setObjectName("horizontalLayout")
        self.pushButton_2 = QtWidgets.QPushButton(StartDialog)
        self.pushButton_2.setStyleSheet("background-color: #8c8880;\n"
"color: rgb(255, 255, 255);")
        self.pushButton_2.setObjectName("pushButton_2")
        self.pushButton_2.clicked.connect(self.exit_button_handler)
        self.horizontalLayout.addWidget(self.pushButton_2)
        self.pushButton = QtWidgets.QPushButton(StartDialog)
        self.pushButton.setStyleSheet("background-color: rgb(113, 177, 217);\n"
"color: rgb(255, 255, 255);")
        self.pushButton.setObjectName("pushButton")
        self.pushButton.clicked.connect(self.start_button_handler)
        self.horizontalLayout.addWidget(self.pushButton)
        self.horizontalLayout.setStretch(0, 20)
        self.horizontalLayout.setStretch(1, 20)
        self.verticalLayout_2.addLayout(self.horizontalLayout)
        self.gridLayout.addLayout(self.verticalLayout_2, 0, 0, 1, 1)

        self.retranslateUi(StartDialog)
        QtCore.QMetaObject.connectSlotsByName(StartDialog)

    def retranslateUi(self, StartDialog):
        _translate = QtCore.QCoreApplication.translate
        StartDialog.setWindowTitle(_translate("StartDialog", "Waste Segregator"))
        self.pushButton_2.setText(_translate("StartDialog", "EXIT"))
        self.pushButton.setText(_translate("StartDialog", "START"))
    
    def exit_button_handler(self):
        print("exit button pressed")
        exitDialog = QtWidgets.QDialog()
        ui = Ui_exitDialog()
        ui.setupUi(exitDialog)
        exitDialog.show()
        exitDialog.exec_()
        

    def start_button_handler(self):
        print("start button clicked")
        self.StartDialog = QtWidgets.QDialog()
        self.MainWindow = QtWidgets.QMainWindow()
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self.MainWindow)
        self.MainWindow.show() #how to execute the main window
        self.StartDialog.accept() #returns 2 and closes the dialog #added self

''' Success class and its functions'''
class Ui_SuccessDialog(object):
    def setupUi(self, SuccessDialog):
        SuccessDialog.setObjectName("SuccessDialog")
        SuccessDialog.resize(400, 300)
        icon = QtGui.QIcon()
        icon.addPixmap(QtGui.QPixmap(":/images/logo-check-1.JPG"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        SuccessDialog.setWindowIcon(icon)
        SuccessDialog.setStyleSheet("background-color: #B3E0F2")
        self.verticalLayout_2 = QtWidgets.QVBoxLayout(SuccessDialog)
        self.verticalLayout_2.setObjectName("verticalLayout_2")
        self.verticalLayout = QtWidgets.QVBoxLayout()
        self.verticalLayout.setObjectName("verticalLayout")
        self.label = QtWidgets.QLabel(SuccessDialog)
        self.label.setStyleSheet("font: 12pt \"Calibri\";\n"
"color: #000000;")
        self.label.setAlignment(QtCore.Qt.AlignCenter)
        self.label.setObjectName("label")
        self.verticalLayout.addWidget(self.label)
        self.pushButton = QtWidgets.QPushButton(SuccessDialog)
        self.pushButton.setStyleSheet("font: 75 12pt \"Calibri\";\n"
"background-color: rgb(113, 177, 217);\n"
"color: rgb(255, 255, 255);")
        self.pushButton.setObjectName("pushButton")
        self.pushButton.clicked.connect(self.ok_button_handler)#############
        self.verticalLayout.addWidget(self.pushButton)
        self.verticalLayout_2.addLayout(self.verticalLayout)

        self.retranslateUi(SuccessDialog)
        QtCore.QMetaObject.connectSlotsByName(SuccessDialog)

    def retranslateUi(self, SuccessDialog):
        _translate = QtCore.QCoreApplication.translate
        SuccessDialog.setWindowTitle(_translate("SuccessDialog", "Waste Segregator"))
        self.label.setText(_translate("SuccessDialog", "Waste segregated successfully!\n"
"Check results section to know more"))
        self.pushButton.setText(_translate("SuccessDialog", "OK"))

    def ok_button_handler(self):
            #app = QtWidgets.QApplication(sys.argv)
            #QtCore.QCoreApplication.instance().quit() #i dont want to quit the application, only the dialog
            SuccessDialog = QtWidgets.QDialog()
            print("Sucess dialog")
            QtWidgets.QDialog().close()
            #SuccessDialog.done(1)
            #SuccessDialog.exit() #not closing
            #takes integer argument'''

''' Exit class and its functions'''


class Ui_exitDialog(object):
    def setupUi(self, exitDialog):
        exitDialog.setObjectName("exitDialog")
        exitDialog.resize(400, 300)
        exitDialog.setStyleSheet("background-color: rgb(179, 224, 242);")
        self.verticalLayout_2 = QtWidgets.QVBoxLayout(exitDialog)
        self.verticalLayout_2.setObjectName("verticalLayout_2")
        self.verticalLayout = QtWidgets.QVBoxLayout()
        self.verticalLayout.setObjectName("verticalLayout")
        self.label = QtWidgets.QLabel(exitDialog)
        self.label.setStyleSheet("font: 12pt \"Calibri\";")
        self.label.setAlignment(QtCore.Qt.AlignCenter)
        self.label.setObjectName("label")
        self.verticalLayout.addWidget(self.label)
        self.horizontalLayout = QtWidgets.QHBoxLayout()
        self.horizontalLayout.setContentsMargins(10, 10, 10, 10)
        self.horizontalLayout.setObjectName("horizontalLayout")
        self.cancelpushButton = QtWidgets.QPushButton(exitDialog)
        self.cancelpushButton.setStyleSheet("background-color: #8c8880;\n"
"color: rgb(255, 255, 255);")
        self.cancelpushButton.setObjectName("cancelpushButton")
        self.cancelpushButton.clicked.connect(self.cancel_button_handler)
        self.horizontalLayout.addWidget(self.cancelpushButton)
        self.exitpushButton = QtWidgets.QPushButton(exitDialog)
        self.exitpushButton.setStyleSheet("background-color: rgb(113, 177, 217);\n"
"color: rgb(255, 255, 255);")
        self.exitpushButton.setObjectName("exitpushButton")
        self.exitpushButton.clicked.connect(self.exit_button_handler)
        self.horizontalLayout.addWidget(self.exitpushButton)
        self.horizontalLayout.setStretch(0, 5)
        self.horizontalLayout.setStretch(1, 5)
        self.verticalLayout.addLayout(self.horizontalLayout)
        self.verticalLayout_2.addLayout(self.verticalLayout)

        self.retranslateUi(exitDialog)
        QtCore.QMetaObject.connectSlotsByName(exitDialog)

    def retranslateUi(self, exitDialog):
        _translate = QtCore.QCoreApplication.translate
        exitDialog.setWindowTitle(_translate("exitDialog", "Dialog"))
        self.label.setText(_translate("exitDialog", "Are you sure you want\n"
" to exit the Application?"))
        self.cancelpushButton.setText(_translate("exitDialog", "Cancel"))
        self.exitpushButton.setText(_translate("exitDialog", "EXIT"))

    def cancel_button_handler(self):
            print("cancel button pressed")
            exitDialog = QtWidgets.QDialog()
            exitDialog.reject()


    def exit_button_handler(self):
            import sys
            app = QtWidgets.QApplication(sys.argv)
            print("exit button pressed")
            exitDialog = QtWidgets.QDialog()
            exitDialog.accept()
            sys.exit(app.quit)



if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    StartDialog = QtWidgets.QDialog() #for start dialog
    ui = Ui_MainWindow()

    uis = Ui_StartDialog() #object of the start class
    uis.setupUi(StartDialog)
    StartDialog.show()
    sys.exit(app.exec_())